from openpyxl.workbook import Workbook
from openpyxl.cell import Cell
from openpyxl import load_workbook
from pydantic import BaseModel

wb: Workbook = load_workbook(filename='utils/dataset.xlsx')
ws = wb['Реестр домов']

headers = ('№ - ТОП', '№', 'Адрес', 'Площадь м2', 'Год', 'Этажей', 'Подьездов', 'Помещений / Квартир')

sh = wb.active

class DatasetRow(BaseModel):
    address: str
    city: str | None
    size: float | None
    population: int | None


def get_data(items: int = 5, page: int = 1) -> list[DatasetRow]:
    data: list[DatasetRow] = []

    row: Cell
    for row in sh.iter_rows():
        # min_row=(page - 1) * items or 2, max_row=page * items + 2
        try:
            address = ''.join((row[2].value).split(',')[:-1])
        except Exception:
            continue
        try:
            city = str(row[2].value).split(',')[-1].strip()
        except Exception:
            continue
        try:
            size = float(row[3].value)
        except Exception:
            continue
        try:
            population = int(row[7].value) * 3
        except Exception:
            population = int(size // 500)
        data.append(DatasetRow(address=address, city=city, size=size, population=population))
        
    return data


if __name__ == "__main__":
    try:
        print(len(get_data()))
    except KeyboardInterrupt as _:
        print('stopped')
